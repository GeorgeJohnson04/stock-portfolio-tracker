import requests
import openpyxl
import os
import sys
import time
from datetime import datetime

FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Finance Project.xlsx")

# Map of ticker symbols to CoinGecko IDs for crypto
CRYPTO_IDS = {
    "BTC": "bitcoin",
    "ETH": "ethereum",
    "SOL": "solana",
    "DOGE": "dogecoin",
    "ADA": "cardano",
    "XRP": "ripple",
    "DOT": "polkadot",
    "AVAX": "avalanche-2",
    "MATIC": "matic-network",
    "LINK": "chainlink",
    "SHIB": "shiba-inu",
    "LTC": "litecoin",
}

def get_crypto_prices(tickers):
    ids = []
    ticker_to_id = {}
    for t in tickers:
        t_upper = t.upper()
        if t_upper in CRYPTO_IDS:
            cg_id = CRYPTO_IDS[t_upper]
            ids.append(cg_id)
            ticker_to_id[cg_id] = t_upper
    if not ids:
        return {}
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {"ids": ",".join(ids), "vs_currencies": "usd"}
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        prices = {}
        for cg_id, t_upper in ticker_to_id.items():
            if cg_id in data and "usd" in data[cg_id]:
                prices[t_upper] = data[cg_id]["usd"]
        return prices
    except Exception as e:
        print(f"  Error fetching crypto prices: {e}")
        return {}

def get_stock_prices(tickers):
    prices = {}
    for ticker in tickers:
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
            headers = {"User-Agent": "Mozilla/5.0"}
            resp = requests.get(url, headers=headers, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            price = data["chart"]["result"][0]["meta"]["regularMarketPrice"]
            prices[ticker.upper()] = price
        except Exception as e:
            print(f"  Error fetching {ticker}: {e}")
    return prices

def refresh():
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Refreshing prices...")

    # Retry if file is locked (e.g. Excel has it open)
    for attempt in range(3):
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
            break
        except PermissionError:
            if attempt < 2:
                print(f"  File locked, retrying in 5s... (attempt {attempt+1}/3)")
                time.sleep(5)
            else:
                print("  ERROR: File is locked. Close Excel and try again.")
                return False

    ws = wb["Holdings"]

    crypto_tickers = []
    stock_tickers = []

    # Scan holdings to find what we need to fetch
    for row in range(4, 56):
        ticker = ws.cell(row=row, column=3).value
        asset_type = ws.cell(row=row, column=2).value
        if not ticker:
            continue
        if asset_type and asset_type.lower() == "crypto":
            crypto_tickers.append(ticker.upper())
        elif asset_type and asset_type.lower() in ("stock", "etf"):
            stock_tickers.append(ticker.upper())

    # Fetch prices
    all_prices = {}
    if crypto_tickers:
        print(f"  Fetching crypto: {', '.join(crypto_tickers)}")
        all_prices.update(get_crypto_prices(crypto_tickers))
    if stock_tickers:
        print(f"  Fetching stocks: {', '.join(stock_tickers)}")
        all_prices.update(get_stock_prices(stock_tickers))

    if not all_prices:
        print("  No prices fetched.")
        wb.close()
        return False

    # Update the Current Price column (column 7 = G)
    updated = 0
    for row in range(4, 56):
        ticker = ws.cell(row=row, column=3).value
        if not ticker:
            continue
        ticker_upper = ticker.upper()
        if ticker_upper in all_prices:
            ws.cell(row=row, column=7).value = all_prices[ticker_upper]
            updated += 1
            print(f"  {ticker_upper}: ${all_prices[ticker_upper]:,.2f}")

    # Save
    try:
        wb.save(FILE_PATH)
        print(f"  Updated {updated} prices. Saved successfully.")
        return True
    except PermissionError:
        print("  ERROR: Cannot save - file is locked. Close Excel and try again.")
        return False
    finally:
        wb.close()

if __name__ == "__main__":
    refresh()
